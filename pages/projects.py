"""Projects view — project list, cost tracking, Excel export."""
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt6.QtWidgets import (QLabel, QPushButton, QComboBox, QLineEdit, QVBoxLayout, 
                             QHBoxLayout, QFrame, QMessageBox, QDoubleSpinBox, QFormLayout, QDialog, QMenu, QFileDialog)
from PyQt6.QtCore import Qt
import config
from helpers import (clear_content, create_table_with_scrollbar, add_row_to_table, sort_table_by_column,
                     BlankDateEdit, apply_saved_sort, create_filter_panel, add_filter_row,
                     create_input_table, add_input_row, add_button_row, create_details_table,
                     copy_table_row_to_clipboard, populate_table, db_fetch_all, db_fetch_scalar, db_execute, confirm_dialog)
from pages import invoices



def get_projects_from_db():
    return db_fetch_all("SELECT * FROM projects ORDER BY start_date DESC")


def get_project_materials_cost(project_id):
    """Get materials cost from invoice items and manual costs for a project."""
    invoice_cost = db_fetch_scalar("""
        SELECT COALESCE(SUM(ii.unit_price * pa.quantity_assigned), 0) as total_cost
        FROM project_assignments pa
        JOIN invoice_items ii ON pa.invoice_item_id = ii.id
        WHERE pa.project_id = ?
    """, (project_id,)) or 0
    
    manual_cost = db_fetch_scalar("""
        SELECT COALESCE(SUM(suma), 0) as total_manual_cost
        FROM manual_costs
        WHERE project_id = ?
    """, (project_id,)) or 0
    
    return invoice_cost + manual_cost


def get_project_labor_cost(project_id):
    """Get labor cost for a project."""
    return db_fetch_scalar("""
        SELECT COALESCE(SUM(total_cost), 0) as total_labor_cost
        FROM labor
        WHERE project_id = ?
    """, (project_id,)) or 0


def get_project_manual_cost(project_id):
    """Get manual costs for a project."""
    return db_fetch_scalar("""
        SELECT COALESCE(SUM(suma), 0) as total_manual_cost
        FROM manual_costs
        WHERE project_id = ?
    """, (project_id,)) or 0


def get_project_assigned_cost(project_id):
    """Get total assigned cost for a project (materials + labor)."""
    return get_project_materials_cost(project_id) + get_project_labor_cost(project_id)


def get_project_statuses():
    return ["W trakcie", "Zakończony", "W planach"]


def _format_project_row(proj):
    """Format a project dict into a row of strings for the table."""
    materials_cost = get_project_materials_cost(proj['id'])
    labor_cost = get_project_labor_cost(proj['id'])
    total_cost = materials_cost + labor_cost
    profit = proj.get('estimated_income', 0) - total_cost
    return [
        proj.get("project_code", ""),
        proj.get("name", ""),
        proj.get("start_date", ""),
        proj.get("end_date") or "",
        proj.get("status", ""),
        f"{proj.get('estimated_income', 0):.2f}",
        f"{materials_cost:.2f}",
        f"{labor_cost:.2f}",
        f"{total_cost:.2f}",
        f"{profit:.2f}"
    ]


def show_projects(window, status_filter=None):
    """Display projects from DB in a sortable table.
    
    Args:
        window: Main window
        status_filter: Optional status to auto-filter ("W planach", "W trakcie", "Zakończony")
    """
    clear_content(window)
    
    # Main horizontal layout for filters (left) and table (right)
    main_h_layout = QHBoxLayout()
    
    # Filter panel (LEFT SIDE)
    filter_frame, filter_table, filter_layout, filter_btn, reset_btn = create_filter_panel(window, 2)
    
    status_combo = QComboBox()
    status_combo.addItems([""] + get_project_statuses())
    add_filter_row(filter_table, 0, "Status:", status_combo, widget_height=27)
    
    year_entry = QLineEdit()
    add_filter_row(filter_table, 1, "Rok:", year_entry, widget_height=27)
    
    # Add button
    add_btn = QPushButton("Dodaj projekt")
    filter_layout.addWidget(add_btn)
    
    filter_layout.addStretch()
    
    main_h_layout.addWidget(filter_frame, 0, Qt.AlignmentFlag.AlignTop)
    
    # Right side layout (RIGHT SIDE) - containing title and table
    right_layout = QVBoxLayout()
    
    # Table (RIGHT SIDE)
    columns = ["Kod", "Nazwa", "Start", "Koniec", "Status", "Przychód", "Koszt materiałów", "Koszt robocizny", "Suma kosztów", "Zysk"]
    table = create_table_with_scrollbar(window, columns, True, "projects_list")
    
    # Initialize sort attributes
    table.sort_column = -1
    table.sort_ascending = True
    
    # Make headers clickable for sorting
    header = table.horizontalHeader()
    header.sectionClicked.connect(lambda col: sort_table_by_column(table, col, numeric_columns=[5, 6, 7, 8, 9], table_name="projects_list"))
    
    right_layout.addWidget(table)
    main_h_layout.addLayout(right_layout)
    
    # Add the main horizontal layout to content
    window.content_layout.addLayout(main_h_layout)
    
    projects_list = get_projects_from_db()
    populate_table(table, projects_list, _format_project_row, numeric_columns=[5, 6, 7, 8, 9])
    
    def filter_action():
        filtered = [
            proj for proj in projects_list
            if (not status_combo.currentText() or proj.get("status") == status_combo.currentText())
            and (not year_entry.text() or str(proj.get("start_date", "")[:4]) == year_entry.text())
        ]
        populate_table(table, filtered, _format_project_row, numeric_columns=[5, 6, 7, 8, 9])
        apply_saved_sort(table, "projects_list", numeric_columns=[5, 6, 7, 8, 9])
    
    def reset_action():
        status_combo.setCurrentIndex(0)
        year_entry.clear()
        populate_table(table, projects_list, _format_project_row, numeric_columns=[5, 6, 7, 8, 9])
        apply_saved_sort(table, "projects_list", numeric_columns=[5, 6, 7, 8, 9])
    
    filter_btn.clicked.connect(filter_action)
    reset_btn.clicked.connect(reset_action)
    add_btn.clicked.connect(lambda: add_project(window))
    
    # Apply status_filter parameter if provided
    if status_filter:
        status_combo.setCurrentText(status_filter)
        filter_action()
    
    # Apply saved sort preference AFTER filtering is done
    apply_saved_sort(table, "projects_list", numeric_columns=[5, 6, 7, 8, 9])
    
    # Context menu for right-clicking on projects
    def show_project_context_menu(pos):
        """Show context menu on right-click for project row."""
        index = table.indexAt(pos)
        if index.row() < 0:
            return
        
        # Get project code from the row
        project_code = table.item(index.row(), 0).text()
        
        # Find the project in the list
        selected_project = None
        for proj in projects_list:
            if proj.get("project_code") == project_code:
                selected_project = proj
                break
        
        if not selected_project:
            return
        
        menu = QMenu()
        
        # Open action
        open_action = menu.addAction("Otwórz")
        open_action.triggered.connect(lambda: show_project_details(window, selected_project.get('project_code')))
        
        # Edit action
        edit_action = menu.addAction("Edytuj")
        edit_action.triggered.connect(lambda: edit_project(window, selected_project))
        
        # Status submenu
        status_menu = menu.addMenu("Zmień status")
        for status in get_project_statuses():
            status_action = status_menu.addAction(status)
            status_action.triggered.connect(lambda checked=False, s=status, p=selected_project: update_project_status(window, p, s))
        
        # Copy action
        copy_action = menu.addAction("Kopiuj")
        copy_action.triggered.connect(lambda: copy_table_row_to_clipboard(table, index.row()))
        
        # Delete action
        delete_action = menu.addAction("Usuń")
        def delete_project():
            if confirm_dialog(window, "Potwierdź", "Czy na pewno chcesz usunąć ten projekt?"):
                try:
                    from helpers import db_execute_many
                    db_execute_many([
                        ("DELETE FROM project_assignments WHERE project_id = ?", (selected_project.get('id'),)),
                        ("DELETE FROM manual_costs WHERE project_id = ?", (selected_project.get('id'),)),
                        ("DELETE FROM projects WHERE id = ?", (selected_project.get('id'),))
                    ])
                    
                    QMessageBox.information(window, "Sukces", "Projekt został usunięty")
                    show_projects(window)
                except Exception as e:
                    QMessageBox.critical(window, "Błąd", f"Błąd przy usuwaniu projektu: {str(e)}")
        delete_action.triggered.connect(delete_project)
        
        menu.exec(table.mapToGlobal(pos))
    
    table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    table.customContextMenuRequested.connect(show_project_context_menu)
    
    # Double-click to show project details
    def on_project_double_click(index):
        if index.row() >= 0:
            project_code = table.item(index.row(), 0).text()
            show_project_details(window, project_code)
    
    table.doubleClicked.connect(on_project_double_click)


def edit_project(window, project):
    """Open dialog to edit an existing project."""
    from PyQt6.QtCore import QDate
    from helpers import BlankDateEdit
    
    dialog = QDialog(window)
    dialog.setWindowTitle("Edytuj projekt")
    dialog.setGeometry(100, 100, 400, 240)
    
    form_layout = QFormLayout()
    
    # Project code (read-only)
    code_input = QLineEdit()
    code_input.setText(project.get("project_code", ""))
    code_input.setReadOnly(True)
    form_layout.addRow("Kod projektu:", code_input)
    
    # Project name
    name_input = QLineEdit()
    name_input.setText(project.get("name", ""))
    form_layout.addRow("Nazwa projektu:", name_input)
    
    # Start date
    start_date = BlankDateEdit()
    if project.get("start_date"):
        start_date.setDate(QDate.fromString(project.get("start_date"), "dd-MM-yyyy"))
    form_layout.addRow("Data startu:", start_date)
    
    # End date (optional)
    end_date = BlankDateEdit()
    if project.get("end_date"):
        end_date.setDate(QDate.fromString(project.get("end_date"), "dd-MM-yyyy"))
    form_layout.addRow("Data końca:", end_date)
    
    # Status
    status_combo = QComboBox()
    status_combo.addItems(get_project_statuses())
    if project.get("status"):
        status_combo.setCurrentText(project.get("status"))
    form_layout.addRow("Status:", status_combo)
    
    # Estimated income
    income_input = QDoubleSpinBox()
    income_input.setMinimum(0)
    income_input.setMaximum(999999)
    income_input.setDecimals(2)
    income_input.setValue(project.get("estimated_income", 0))
    form_layout.addRow("Przychód szacunkowy:", income_input)
    
    # Buttons
    button_layout = QHBoxLayout()
    save_btn = QPushButton("Zapisz")
    button_layout.addWidget(save_btn)
    form_layout.addRow(button_layout)
    
    dialog.setLayout(form_layout)
    
    def save_project():
        if not name_input.text():
            QMessageBox.warning(dialog, "Błąd", "Nazwa projektu jest wymagana")
            return
        
        try:
            end_date_val = end_date.date().toString("dd-MM-yyyy") if end_date.date() != end_date.minimumDate() else None
            start_date_val = start_date.date().toString("dd-MM-yyyy") if start_date.date() != start_date.minimumDate() else ""
            
            db_execute("""
                UPDATE projects 
                SET name = ?, start_date = ?, end_date = ?, status = ?, estimated_income = ?
                WHERE project_code = ?
            """, (name_input.text(), start_date_val, end_date_val, 
                  status_combo.currentText(), income_input.value(), project.get("project_code")))
            
            QMessageBox.information(dialog, "Sukces", "Projekt został zaktualizowany")
            dialog.accept()
            show_projects(window)
        except Exception as e:
            QMessageBox.critical(dialog, "Błąd", f"Nie można zaktualizować projektu: {str(e)}")
    
    save_btn.clicked.connect(save_project)
    
    dialog.exec()


def update_project_status(window, project, status):
    """Update project status."""
    try:
        db_execute("""
            UPDATE projects 
            SET status = ?
            WHERE project_code = ?
        """, (status, project.get("project_code")))
        
        show_projects(window)
    except Exception as e:
        QMessageBox.critical(window, "Błąd", f"Nie można zaktualizować statusu: {str(e)}")

def add_project(window):
    """Open dialog to add a new project."""
    from PyQt6.QtWidgets import QDialog, QSpinBox, QDoubleSpinBox
    from PyQt6.QtCore import QDate
    from helpers import BlankDateEdit
    
    dialog = QDialog(window)
    dialog.setWindowTitle("Dodaj projekt")
    dialog.setGeometry(100, 100, 530, 285)
    
    layout = QVBoxLayout()
    
    # Create table for inputs
    input_table = create_input_table(7)
    
    # Row 0: Project code
    code_input = QLineEdit()
    add_input_row(input_table, 0, "Kod projektu:", code_input)
    
    # Row 1: Project name
    name_input = QLineEdit()
    add_input_row(input_table, 1, "Nazwa projektu:", name_input)
    
    # Row 2: Start date
    start_date = BlankDateEdit()
    add_input_row(input_table, 2, "Data startu:", start_date)
    
    # Row 3: End date
    end_date = BlankDateEdit()
    add_input_row(input_table, 3, "Data końca:", end_date)
    
    # Row 4: Status
    status_combo = QComboBox()
    status_combo.addItems(get_project_statuses())
    add_input_row(input_table, 4, "Status:", status_combo)
    
    # Row 5: Estimated income
    income_input = QDoubleSpinBox()
    income_input.setMinimum(0)
    income_input.setMaximum(999999)
    income_input.setDecimals(2)
    add_input_row(input_table, 5, "Przychód szacunkowy:", income_input)
    
    # Row 6: Buttons
    save_btn = QPushButton("Zapisz")
    save_btn.setDefault(True)
    add_button_row(input_table, 6, save_btn)
    
    layout.addWidget(input_table)
    dialog.setLayout(layout)
    
    def save_project():
        if not code_input.text() or not name_input.text():
            QMessageBox.warning(dialog, "Błąd", "Kod i nazwa projektu są wymagane")
            return
        
        try:
            end_date_val = end_date.date().toString("dd-MM-yyyy") if end_date.date() != end_date.minimumDate() else None
            start_date_val = start_date.date().toString("dd-MM-yyyy") if start_date.date() != start_date.minimumDate() else ""
            
            db_execute("""
                INSERT INTO projects (project_code, name, start_date, end_date, status, estimated_income)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (code_input.text(), name_input.text(), start_date_val, end_date_val, 
                  status_combo.currentText(), income_input.value()))
            
            dialog.accept()
            show_projects(window)
        except Exception as e:
            QMessageBox.critical(dialog, "Błąd", f"Nie można dodać projektu: {str(e)}")
    
    save_btn.clicked.connect(save_project)
    
    dialog.exec()


def get_project_details_from_db(project_code):
    """Get project details and all items assigned to it."""
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get project
    c.execute("SELECT * FROM projects WHERE project_code = ?", (project_code,))
    project_row = c.fetchone()
    project = dict(project_row) if project_row else None
    
    if not project:
        conn.close()
        return None
    
    project_id = project['id']
    
    # Get assigned items
    c.execute("""
        SELECT ii.*, pa.quantity_assigned, i.invoice_number, i.seller_name, i.invoice_date
        FROM project_assignments pa
        JOIN invoice_items ii ON pa.invoice_item_id = ii.id
        JOIN invoices i ON ii.invoice_id = i.id
        WHERE pa.project_id = ?
        ORDER BY i.invoice_date DESC
    """, (project_id,))
    assigned_items = [dict(row) for row in c.fetchall()]
    
    # Get labor entries
    c.execute("""
        SELECT l.*, e.name as person_name
        FROM labor l
        LEFT JOIN employees e ON l.employee_id = e.id
        WHERE l.project_id = ?
        ORDER BY l.work_date DESC
    """, (project_id,))
    labor_items = [dict(row) for row in c.fetchall()]
    
    # Get manual costs
    c.execute("""
        SELECT * FROM manual_costs
        WHERE project_id = ?
        ORDER BY data DESC
    """, (project_id,))
    manual_items = [dict(row) for row in c.fetchall()]
    
    conn.close()
    
    return {
        "project": project,
        "assigned_items": assigned_items,
        "labor_items": labor_items,
        "manual_items": manual_items
    }


def export_project_materials_to_xlsx(window, project_code, project_name, items_table):
    """Export project materials table to XLSX file."""
    try:
        # Get file path from user
        file_path, _ = QFileDialog.getSaveFileName(
            window,
            "Eksportuj materiały do Excel",
            f"Projekt_{project_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Materiały"
        
        # Add title
        title_cell = ws['A1']
        title_cell.value = f"Materiały projektu: {project_name}"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        # Add date
        date_str = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        date_cell = ws['A2']
        date_cell.value = f"Data eksportu: {date_str}"
        date_cell.font = Font(italic=True, size=10)
        ws.row_dimensions[2].height = 20
        
        # Add headers in row 4
        headers = []
        for col in range(items_table.columnCount()):
            header_item = items_table.horizontalHeaderItem(col)
            if header_item:
                headers.append(header_item.text())
            else:
                headers.append("")
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.row_dimensions[4].height = 25
        
        # Add data rows
        numeric_columns = [5, 6, 7]  # Ilość, Cena jednostkowa, Suma
        
        for row_idx in range(items_table.rowCount()):
            for col_idx in range(items_table.columnCount()):
                item = items_table.item(row_idx, col_idx)
                cell = ws.cell(row=row_idx + 5, column=col_idx + 1)
                
                if item:
                    cell.value = item.text()
                    
                    # Format numeric columns
                    if col_idx + 1 in numeric_columns:
                        try:
                            # Convert Polish decimal separator to standard
                            value_str = item.text().replace(',', '.')
                            cell.value = float(value_str)
                            cell.number_format = '0.00'
                        except ValueError:
                            cell.value = item.text()
                    
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = border
            
            ws.row_dimensions[row_idx + 5].height = 20
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Dokument
        ws.column_dimensions['B'].width = 20  # Firma
        ws.column_dimensions['C'].width = 12  # Data
        ws.column_dimensions['D'].width = 25  # Opis pozycji
        ws.column_dimensions['E'].width = 12  # Ilość
        ws.column_dimensions['F'].width = 18  # Cena jednostkowa
        ws.column_dimensions['G'].width = 15  # Suma
        ws.column_dimensions['H'].width = 15  # Ręcznie dodane
        
        # Save file
        wb.save(file_path)
        
        QMessageBox.information(window, "Sukces", f"Materiały zostały wyeksportowane do:\n{file_path}")
        
    except Exception as e:
        QMessageBox.critical(window, "Błąd", f"Błąd przy eksporcie: {str(e)}")


def manual_cost_dialog(project_id, window, project_code, manual_cost_id=None):
    """Show dialog to add or edit manual cost for a project."""
    from PyQt6.QtWidgets import QDialog
    from PyQt6.QtCore import QDate
    from helpers import db_fetch_one
    
    edit_mode = manual_cost_id is not None
    current_data = {}
    
    if edit_mode:
        current_data = db_fetch_one("SELECT * FROM manual_costs WHERE id = ?", (manual_cost_id,))
        if not current_data:
            QMessageBox.warning(window, "Błąd", "Nie znaleziono kosztu.")
            return
    
    dialog = QDialog(window)
    dialog.setWindowTitle("Edytuj koszt ręcznie" if edit_mode else "Dodaj koszt ręcznie")
    dialog.setGeometry(200, 200, 500, 317)
    
    layout = QVBoxLayout(dialog)
    
    # Create table for inputs
    input_table = create_input_table(8, col2_width=320)
    
    # Row 0: Dokument
    dok_input = QLineEdit()
    if edit_mode:
        dok_input.setText(current_data.get('dokument', ''))
    add_input_row(input_table, 0, "Dokument:", dok_input)
    
    # Row 1: Firma
    firma_input = QLineEdit()
    if edit_mode:
        firma_input.setText(current_data.get('firma', ''))
    add_input_row(input_table, 1, "Firma:", firma_input)
    
    # Row 2: Data
    data_input = BlankDateEdit()
    if edit_mode:
        date_str = current_data.get('data', '')
        if date_str:
            date_parts = date_str.split('-')
            if len(date_parts) == 3:
                data_input.setDate(QDate(int(date_parts[2]), int(date_parts[1]), int(date_parts[0])))
        else:
            data_input.setDate(QDate.currentDate())
    else:
        data_input.setDate(QDate.currentDate())
    add_input_row(input_table, 2, "Data:", data_input)
    
    # Row 3: Opis pozycji
    opis_input = QLineEdit()
    if edit_mode:
        opis_input.setText(current_data.get('opis_pozycji', ''))
    add_input_row(input_table, 3, "Opis pozycji:", opis_input)
    
    # Row 4: Ilość
    ilosc_input = QLineEdit()
    if edit_mode:
        ilosc_input.setText(f"{current_data.get('ilosc', 0):.2f}".replace('.', ','))
    add_input_row(input_table, 4, "Ilość:", ilosc_input)
    
    # Row 5: Cena jednostkowa
    cena_input = QLineEdit()
    if edit_mode:
        cena_input.setText(f"{current_data.get('cena_jednostkowa', 0):.2f}".replace('.', ','))
    add_input_row(input_table, 5, "Cena jednostkowa:", cena_input)
    
    # Row 6: Suma (read-only)
    suma_input = QLineEdit()
    suma_input.setReadOnly(True)
    add_input_row(input_table, 6, "Suma:", suma_input)
    
    # Function to calculate suma
    def update_suma():
        try:
            ilosc_str = ilosc_input.text().strip().replace(',', '.')
            cena_str = cena_input.text().strip().replace(',', '.')
            ilosc = float(ilosc_str) if ilosc_str else 0
            cena = float(cena_str) if cena_str else 0
            suma = ilosc * cena
            suma_input.setText(f"{suma:.2f}".replace('.', ','))
        except:
            suma_input.setText("0,00")
    
    ilosc_input.textChanged.connect(update_suma)
    cena_input.textChanged.connect(update_suma)
    if edit_mode:
        update_suma()
    
    # Row 7: Buttons
    ok_btn = QPushButton("Zapisz")
    ok_btn.setDefault(True)
    add_button_row(input_table, 7, ok_btn)
    
    layout.addWidget(input_table)
    dialog.setLayout(layout)
    
    ok_btn.clicked.connect(dialog.accept)
    
    if dialog.exec() == QDialog.DialogCode.Accepted:
        try:
            dokument = dok_input.text().strip()
            firma = firma_input.text().strip()
            data = data_input.date().toString("dd-MM-yyyy")
            opis_pozycji = opis_input.text().strip()
            ilosc_str = ilosc_input.text().strip().replace(',', '.')
            cena_str = cena_input.text().strip().replace(',', '.')
            ilosc = float(ilosc_str) if ilosc_str else 0
            cena = float(cena_str) if cena_str else 0
            suma = ilosc * cena
            
            if ilosc <= 0 or cena <= 0:
                QMessageBox.warning(dialog, "Błąd", "Ilość i cena muszą być większe od 0.")
                return False
            
            if edit_mode:
                db_execute("""
                    UPDATE manual_costs
                    SET dokument=?, firma=?, data=?, opis_pozycji=?, ilosc=?, cena_jednostkowa=?, suma=?
                    WHERE id = ?
                """, (dokument, firma, data, opis_pozycji, ilosc, cena, suma, manual_cost_id))
                QMessageBox.information(dialog, "Sukces", "Koszt zaktualizowany pomyślnie.")
            else:
                db_execute("""
                    INSERT INTO manual_costs (project_id, dokument, firma, data, opis_pozycji, ilosc, cena_jednostkowa, suma)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (project_id, dokument, firma, data, opis_pozycji, ilosc, cena, suma))
                QMessageBox.information(dialog, "Sukces", "Koszt dodany pomyślnie.")
            
            show_project_details(window, project_code)
            return True
            
        except ValueError:
            QMessageBox.warning(dialog, "Błąd", "Niepoprawny format danych.")
            return False
    
    return False


def show_project_details(window, project_code):
    """Show project details with assigned items."""
    clear_content(window)
    
    # Get project and items from database
    project_data = get_project_details_from_db(project_code)
    
    if not project_data:
        error_label = QLabel("Projekt nie znaleziony")
        window.content_layout.addWidget(error_label)
        return
    
    project = project_data["project"]
    assigned_items = project_data["assigned_items"]
    labor_items = project_data["labor_items"]
    manual_items = project_data["manual_items"]
    
    # Calculate total cost (items + labor + manual) and profit
    items_cost = sum(item['unit_price'] * item['quantity_assigned'] for item in assigned_items) if assigned_items else 0
    labor_cost = sum(item.get('total_cost', 0) for item in labor_items) if labor_items else 0
    manual_cost = sum(item.get('suma', 0) for item in manual_items) if manual_items else 0
    total_cost = items_cost + labor_cost + manual_cost
    profit = project.get('estimated_income', 0) - total_cost
    
    # Main horizontal layout
    main_h_layout = QHBoxLayout()
    
    # LEFT SIDE - Project Details Frame
    details_frame = QFrame(window)
    details_frame.setMaximumWidth(275)
    details_layout = QVBoxLayout(details_frame)
    details_layout.setContentsMargins(10, 10, 10, 10)
    
    # Format dates
    start_date = project.get('start_date', '') or ''
    end_date = project.get('end_date', '') or ''
    
    # Project details table - 2 columns (label, value) with all main list info
    detail_items = [
        ("Kod:", project.get('project_code', '')),
        ("Nazwa:", project.get('name', '')),
        ("Start:", start_date),
        ("Koniec:", end_date),
        ("Status:", project.get('status', '')),
        ("Przychód:", f"{project.get('estimated_income', 0):.2f}".replace('.', ',')),
        ("Koszt materiałów:", f"{items_cost:.2f}".replace('.', ',')),
        ("Koszt robocizny:", f"{labor_cost:.2f}".replace('.', ',')),
        ("Suma kosztów:", f"{total_cost:.2f}".replace('.', ',')),
        ("Zysk:", f"{profit:.2f}".replace('.', ','))
    ]
    
    details_table = create_details_table(detail_items, row_height=31)
    
    details_layout.addWidget(details_table)
    
    # Back button
    back_btn = QPushButton("Powrót do projektów")
    back_btn.clicked.connect(lambda: show_projects(window))
    details_layout.addWidget(back_btn)
    
    # Add manual cost button
    add_cost_btn = QPushButton("Dodaj koszt ręcznie")
    add_cost_btn.clicked.connect(lambda: manual_cost_dialog(project.get('id'), window, project_code))
    details_layout.addWidget(add_cost_btn)
    
    # Export button
    export_btn = QPushButton("Eksportuj do Excel")
    export_btn.clicked.connect(lambda: export_project_materials_to_xlsx(window, project.get('project_code', ''), project.get('name', ''), items_table))
    details_layout.addWidget(export_btn)
    
    # Delete project button
    delete_btn = QPushButton("Usuń projekt")
    def delete_project_from_details():
        if confirm_dialog(window, "Potwierdź", "Czy na pewno chcesz usunąć ten projekt?"):
            try:
                conn = sqlite3.connect(config.DB_PATH)
                c = conn.cursor()
                
                # Delete from project_assignments
                c.execute("""DELETE FROM project_assignments 
                           WHERE project_id = ?""", (project.get('id'),))
                
                # Delete from manual_costs
                c.execute("""DELETE FROM manual_costs 
                           WHERE project_id = ?""", (project.get('id'),))
                
                # Delete the project
                c.execute("DELETE FROM projects WHERE id = ?", (project.get('id'),))
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(window, "Sukces", "Projekt został usunięty")
                show_projects(window)
            except Exception as e:
                QMessageBox.critical(window, "Błąd", f"Błąd przy usuwaniu projektu: {str(e)}")
    delete_btn.clicked.connect(delete_project_from_details)
    details_layout.addWidget(delete_btn)
    
    details_layout.addStretch()
    
    main_h_layout.addWidget(details_frame)
    
    # RIGHT SIDE - Assigned Items Table
    right_layout = QVBoxLayout()
    
    # Items table
    columns = ["Dokument", "Firma", "Data", "Opis pozycji", "Ilość", "Cena jednostkowa", "Suma", "Ręcznie dodane"]
    items_table = create_table_with_scrollbar(window, columns, True, "project_details_items")
    
    # Track manual items for context menu
    manual_items_indices = []
    
    # Populate items table with invoice items
    for item in assigned_items:
        item_total = item['unit_price'] * item['quantity_assigned']
        add_row_to_table(items_table, [
            item.get('invoice_number', ''),
            item.get('seller_name', ''),
            item.get('invoice_date', ''),
            item.get('item_description', ''),
            f"{item['quantity_assigned']:.2f}".replace('.', ','),
            f"{item['unit_price']:.2f}".replace('.', ','),
            f"{item_total:.2f}".replace('.', ','),
            ""
        ], numeric_columns=[4, 5, 6])
    
    # Track where invoice items end (to distinguish from labor and manual items)
    invoice_items_count = len(assigned_items)
    
    # Populate items table with labor items
    for labor in labor_items:
        labor_total = labor.get('total_cost', 0)
        add_row_to_table(items_table, [
            "Lista płac",
            "Własna",
            labor.get('work_date', ''),
            f"Robocizna: {labor.get('person_name', '')}",
            f"{labor.get('hours_worked', 0):.2f}".replace('.', ','),
            f"{labor.get('hourly_rate', 0):.2f}".replace('.', ','),
            f"{labor_total:.2f}".replace('.', ','),
            ""
        ], numeric_columns=[4, 5, 6])
    
    # Populate items table with manual costs
    for idx, manual in enumerate(manual_items):
        manual_items_indices.append(items_table.rowCount())
        add_row_to_table(items_table, [
            manual.get('dokument', ''),
            manual.get('firma', ''),
            manual.get('data', ''),
            manual.get('opis_pozycji', ''),
            f"{manual.get('ilosc', 0):.2f}".replace('.', ','),
            f"{manual.get('cena_jednostkowa', 0):.2f}".replace('.', ','),
            f"{manual.get('suma', 0):.2f}".replace('.', ','),
            "✓"
        ], numeric_columns=[4, 5, 6])
        
        # Center the checkmark in the last column
        checkmark_item = items_table.item(items_table.rowCount() - 1, 7)
        if checkmark_item:
            checkmark_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
    
    # Add double-click handler for invoice items
    def on_item_double_clicked(index):
        """Handle double-click on items table to open invoice details."""
        row = index.row()
        # Only open invoice details if clicking on invoice items (first rows)
        # Rows after invoice_items_count are labor or manual items
        if row < invoice_items_count:
            dokument_item = items_table.item(row, 0)
            if dokument_item:
                dokument = dokument_item.text()
                if dokument:
                    invoices.show_invoice_details(window, dokument)
    
    items_table.doubleClicked.connect(on_item_double_clicked)
    
    # Add context menu for all items
    def show_item_context_menu(pos):
        """Show context menu for right-click on items."""
        index = items_table.indexAt(pos)
        row = index.row()
        
        menu = QMenu()
        
        # Copy action for all items
        copy_action = menu.addAction("Kopiuj")
        copy_action.triggered.connect(lambda: copy_table_row_to_clipboard(items_table, row))
        
        # Edit and Delete options only for manual items
        if row in manual_items_indices:
            menu.addSeparator()
            edit_action = menu.addAction("Edytuj")
            delete_action = menu.addAction("Usuń")
            
            def edit_item():
                manual_item = manual_items[manual_items_indices.index(row)]
                manual_cost_dialog(project.get('id'), window, project_code, manual_item.get('id'))
            
            def delete_item():
                if confirm_dialog(window, "Potwierdź", "Czy na pewno chcesz usunąć ten koszt?"):
                    manual_item = manual_items[manual_items_indices.index(row)]
                    db_execute("DELETE FROM manual_costs WHERE id = ?", (manual_item.get('id'),))
                    show_project_details(window, project_code)
            
            edit_action.triggered.connect(edit_item)
            delete_action.triggered.connect(delete_item)
        
        menu.exec(items_table.mapToGlobal(pos))
    
    items_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    items_table.customContextMenuRequested.connect(show_item_context_menu)
    
    right_layout.addWidget(items_table)
    main_h_layout.addLayout(right_layout)
    
    # Add the main horizontal layout to content
    window.content_layout.addLayout(main_h_layout)
